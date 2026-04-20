#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl, json, sys, os, re
from datetime import date

XLSX = "통합 문서1.xlsx"

def parse_date(s: str):
    if not s: return None
    m = re.match(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})", s.strip())
    if not m: return None
    y, mo, d = map(int, m.groups())
    return f"{y:04d}-{mo:02d}-{d:02d}"

def classify_type(kind: str):
    k = (kind or "").strip()
    if k in ("연차", "연차(1일)"): return ("annual", 1.0)
    if "오전" in k and ("반반" in k or "반차" in k and "4시간" not in k):
        return ("quarter_am", 0.25)
    if "오후" in k and ("반반" in k or "반차" in k and "4시간" not in k):
        return ("quarter_pm", 0.25)
    if "오전반차" in k: return ("half_am", 0.5)
    if "오후반차" in k: return ("half_pm", 0.5)
    if "반반차오후" in k: return ("quarter_pm", 0.25)
    if "반반차오전" in k: return ("quarter_am", 0.25)
    if "시간차" in k: return ("hourly", None)   # days는 duration으로 계산
    if "병가" in k: return ("family_care", None)
    if "경조" in k: return ("condolence_close", None)
    if "예비군" in k or "민방위" in k or "공가" in k: return ("military", None)
    if "생리" in k: return ("menstrual", 1.0)
    return (k, None)   # unknown — 원본 저장

def parse_duration(s: str, ltype: str):
    """'1일', '4시간', '2시간', '가족필요1일' 등 → days 수치"""
    if not s: return None
    s = s.strip()
    mday = re.search(r"(\d+(?:\.\d+)?)\s*일", s)
    if mday: return float(mday.group(1))
    mhr  = re.search(r"(\d+(?:\.\d+)?)\s*시간", s)
    if mhr:
        hours = float(mhr.group(1))
        # 4시간 = 반차 0.5
        if abs(hours - 4) < 0.1: return 0.5
        if abs(hours - 2) < 0.1: return 0.25
        return round(hours / 8, 3)
    return None

wb = openpyxl.load_workbook(XLSX, data_only=True)
result = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    # 시트 최상단 헤더 스킵, 연속된 4-row 블록으로 취급
    records = []
    i = 0
    while i < len(rows):
        block = rows[i:i+4]
        texts = [str(r[1]) if r and len(r) > 1 and r[1] else "" for r in block]
        # 표준 4-row 패턴: [status, type, date, duration]
        if len(texts) >= 3 and texts[0] and "승인" in texts[0]:
            ltype_raw = texts[1] if len(texts) > 1 else ""
            d = parse_date(texts[2]) if len(texts) > 2 else None
            dur_raw = texts[3] if len(texts) > 3 else ""
            # duration row 없는 3-row 블록도 허용
            has_dur = bool(dur_raw) and not ("승인" in dur_raw)
            if d:
                type_key, default_days = classify_type(ltype_raw)
                days = parse_duration(dur_raw, type_key) or default_days or 1.0
                records.append({
                    "date": d,
                    "type_raw": ltype_raw,
                    "type_key": type_key,
                    "duration_raw": dur_raw if has_dur else "(기본 1일 적용)",
                    "days": days,
                })
            i += 4 if has_dur else 3
        else:
            i += 1
    result[name] = records

# 2025-01-01 ~ 2026-04-13 필터
start = "2025-01-01"
end = "2026-04-13"
filtered = {}
for emp, recs in result.items():
    keep = [r for r in recs if start <= r["date"] <= end]
    if keep:
        filtered[emp] = keep

with open("scripts/.leave-parsed.json", "w", encoding="utf-8") as f:
    json.dump(filtered, f, ensure_ascii=False, indent=2)

total = sum(len(v) for v in filtered.values())
print(f"직원 수: {len(filtered)}")
print(f"총 휴가 건수: {total}")
for emp, recs in filtered.items():
    types = {}
    for r in recs:
        types[r["type_key"]] = types.get(r["type_key"], 0) + r["days"]
    print(f"  {emp}: {len(recs)}건, 타입별={types}")
