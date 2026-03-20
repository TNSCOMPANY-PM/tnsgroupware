"""
엑셀 시트 구조 진단 스크립트
python scripts/inspect-xlsx.py
"""
import openpyxl
import os

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "(주)티앤에스컴퍼니 매출 통계 24.01~ (4).xlsx")

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
print("시트 목록:", wb.sheetnames)

TARGET_SHEETS = ["26년 1월", "26년 2월", "26년 3월"]

for sheet_name in TARGET_SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"\n[{sheet_name}] 시트 없음")
        continue

    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"[{sheet_name}] max_row={ws.max_row}, max_col={ws.max_column}")

    # 헤더 구조 파악: 1~40행을 출력
    print("\n--- 1~40행 구조 ---")
    for row_idx in range(1, 41):
        row_vals = []
        for col_idx in range(1, 25):  # A~X 열
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is not None:
                row_vals.append(f"[{col_idx}]{repr(v)[:30]}")
        if row_vals:
            print(f"  행{row_idx}: {' | '.join(row_vals)}")

    # 합계 행 찾기
    sum_row = None
    header_row = None
    for row_idx in range(1, 50):
        c_val = ws.cell(row=row_idx, column=3).value
        if c_val == "합계":
            sum_row = row_idx
        if c_val == "NO":
            header_row = row_idx
            break

    print(f"\n  합계 행: {sum_row}, 헤더(NO) 행: {header_row}")

    if sum_row:
        print("  합계 행 전체 값:")
        for col_idx in range(1, 25):
            v = ws.cell(row=sum_row, column=col_idx).value
            if v is not None:
                print(f"    col{col_idx}({chr(64+col_idx)}): {repr(v)}")

    if header_row:
        print("  헤더 행 전체 값:")
        for col_idx in range(1, 25):
            v = ws.cell(row=header_row, column=col_idx).value
            if v is not None:
                print(f"    col{col_idx}({chr(64+col_idx)}): {repr(v)}")

    # 데이터 시작 후 첫 10행 샘플
    if header_row:
        data_start = header_row + 1
        print(f"\n  데이터 샘플 (행{data_start}~{data_start+9}):")
        for row_idx in range(data_start, data_start + 10):
            row_vals = []
            for col_idx in range(1, 20):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    row_vals.append(f"[{col_idx}]{repr(v)[:25]}")
            if row_vals:
                print(f"    행{row_idx}: {' | '.join(row_vals)}")

    # 전체 데이터 카운트 및 합계
    if header_row:
        data_start = header_row + 1
        dep_count = 0
        dep_sum = 0
        wit_count = 0
        wit_sum = 0
        for row_idx in range(data_start, ws.max_row + 1):
            dep_date = ws.cell(row=row_idx, column=4).value   # D열 (1-based = col 4)
            dep_amt  = ws.cell(row=row_idx, column=9).value   # I열
            wit_date = ws.cell(row=row_idx, column=11).value  # K열
            wit_amt  = ws.cell(row=row_idx, column=17).value  # Q열

            if dep_date and dep_amt:
                try:
                    dep_sum += round(float(str(dep_amt).replace(",","")))
                    dep_count += 1
                except:
                    pass
            if wit_date and wit_amt:
                try:
                    wit_sum += round(float(str(wit_amt).replace(",","")))
                    wit_count += 1
                except:
                    pass

        # 시트 합계값
        if sum_row:
            sheet_dep = ws.cell(row=sum_row, column=9).value
            sheet_wit = ws.cell(row=sum_row, column=17).value
        else:
            sheet_dep = sheet_wit = "?"

        print(f"\n  파싱 결과: 매출 {dep_count}건 {dep_sum:,}원 | 매입 {wit_count}건 {wit_sum:,}원")
        print(f"  시트합계:  매출 {sheet_dep} | 매입 {sheet_wit}")
