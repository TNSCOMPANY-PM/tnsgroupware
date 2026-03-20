"""
TNS 매출통계 xlsx → Supabase finance 재임포트 (Python + openpyxl)
- 2026-01, 2026-02, 2026-03 기존 데이터를 완전 삭제 (Pushbullet 제외)
- openpyxl로 정확하게 재파싱 후 삽입

python scripts/reimport-xlsx-python.py
"""
import openpyxl
import os, json, hashlib, urllib.request, urllib.parse

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "(주)티앤에스컴퍼니 매출 통계 24.01~ (4).xlsx")
HOST = "REDACTED_PROJECT_REF.supabase.co"
KEY  = "<REDACTED_SERVICE_ROLE_KEY>"
BASE = f"https://{HOST}/rest/v1"

HEADERS = {
    "Content-Type": "application/json",
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Prefer": "resolution=ignore-duplicates,return=minimal",
}

TARGET_MONTHS = ["2026-01", "2026-02", "2026-03"]
SHEET_MAP = {
    "26년 1월": "2026-01",
    "26년 2월": "2026-02",
    "26년 3월": "2026-03",
}

# ── 결정적 UUID (row_num 포함 → 중복 행 구분) ─────────────
def make_uuid(date, typ, amount, desc, row_num):
    raw = f"xlsx2|{date}|{typ}|{amount}|{desc}|{row_num}"
    h = hashlib.md5(raw.encode("utf-8")).hexdigest()
    return f"{h[:8]}-{h[8:12]}-4{h[13:16]}-{h[16:20]}-{h[20:32]}"

# ── HTTP 유틸 ─────────────────────────────────────────────
def api(method, path, body=None, extra_headers=None):
    url = BASE + path
    data = json.dumps(body).encode() if body is not None else None
    headers = dict(HEADERS)
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, r.read().decode()
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()

def fetch_all(path):
    """페이지네이션으로 전체 행 가져오기"""
    rows = []
    offset = 0
    limit = 1000
    while True:
        sep = "&" if "?" in path else "?"
        status, body = api("GET", f"{path}{sep}limit={limit}&offset={offset}",
                           extra_headers={"Range": f"{offset}-{offset+limit-1}"})
        chunk = json.loads(body)
        if not isinstance(chunk, list):
            print("오류:", body[:300])
            break
        rows.extend(chunk)
        if len(chunk) < limit:
            break
        offset += limit
    return rows

# ── xlsx 파싱 ─────────────────────────────────────────────
def parse_sheet(ws, month):
    rows = []

    # 합계/NO 행 탐색
    sum_row = header_row = None
    for r in range(1, 50):
        v = ws.cell(row=r, column=3).value
        if v == "합계":
            sum_row = r
        if v == "NO":
            header_row = r
            break

    if not header_row:
        print(f"[{month}] 헤더 행 없음")
        return rows

    expected_dep = round(float(ws.cell(row=sum_row, column=9).value or 0)) if sum_row else 0
    expected_wit = round(float(ws.cell(row=sum_row, column=17).value or 0)) if sum_row else 0

    data_start = header_row + 1
    actual_dep = actual_wit = 0

    for r in range(data_start, ws.max_row + 1):
        dep_date_raw = ws.cell(row=r, column=4).value
        dep_amt_raw  = ws.cell(row=r, column=9).value
        wit_date_raw = ws.cell(row=r, column=11).value
        wit_amt_raw  = ws.cell(row=r, column=17).value

        # 날짜 변환
        def to_date(v):
            if v is None:
                return None
            if hasattr(v, "strftime"):
                return v.strftime("%Y-%m-%d")
            return None

        dep_date = to_date(dep_date_raw)
        wit_date = to_date(wit_date_raw)

        def to_int(v):
            if v is None:
                return 0
            try:
                return round(float(str(v).replace(",", "")))
            except:
                return 0

        dep_amt = to_int(dep_amt_raw)
        wit_amt = to_int(wit_amt_raw)

        # 매출 행
        if dep_date and dep_amt != 0:
            name    = str(ws.cell(row=r, column=5).value or "").strip()
            team    = str(ws.cell(row=r, column=6).value or "").strip()
            content = str(ws.cell(row=r, column=8).value or "").strip()
            desc    = f"{name} - {content}" if content else name
            rows.append({
                "id":          make_uuid(dep_date, "매출", dep_amt, desc, r),
                "month":       month,
                "date":        dep_date,
                "type":        "매출",
                "amount":      dep_amt,
                "category":    team or None,
                "description": desc,
                "client_name": name,
                "status":      "completed",
            })
            actual_dep += dep_amt

        # 매입/환불 행
        if wit_date and wit_amt != 0:
            name    = str(ws.cell(row=r, column=12).value or "").strip()
            team    = str(ws.cell(row=r, column=13).value or "").strip()
            subtype = str(ws.cell(row=r, column=14).value or "").strip()
            content = str(ws.cell(row=r, column=16).value or "").strip()
            desc    = f"{name} - {content}" if content else name
            rows.append({
                "id":          make_uuid(wit_date, "매입", wit_amt, desc, r),
                "month":       month,
                "date":        wit_date,
                "type":        "매입",
                "amount":      wit_amt,
                "category":    team or None,
                "description": desc,
                "client_name": name,
                "status":      "completed",
            })
            actual_wit += wit_amt

        # 연속 5행 빈 경우 종료
        if not dep_date and not wit_date:
            blank = sum(
                1 for rr in range(r, min(r + 5, ws.max_row + 1))
                if ws.cell(row=rr, column=4).value is None
                and ws.cell(row=rr, column=11).value is None
            )
            if blank >= 5:
                break

    dep_ok = "OK" if actual_dep == expected_dep else f"MISMATCH (sheet:{expected_dep:,})"
    wit_ok = "OK" if actual_wit == expected_wit else f"MISMATCH (sheet:{expected_wit:,})"
    print(f"[{month}] dep {sum(1 for r in rows if r['type']=='매출')} rows {actual_dep:,} {dep_ok} "
          f"| wit {sum(1 for r in rows if r['type']=='매입')} rows {actual_wit:,} {wit_ok}")
    return rows

# ── 메인 ─────────────────────────────────────────────────
def main():
    # Pushbullet 행 전체 필드 백업 (삭제 전에!)
    print("=== Pushbullet 행 백업 ===")
    pb_rows = []
    for month in TARGET_MONTHS:
        rows = fetch_all(f"/finance?month=eq.{month}&description=like.*pb%3Asms%3A*&select=*")
        pb_rows.extend(rows)
    print(f"  Pushbullet rows backed up: {len(pb_rows)}")

    # 월별 전체 삭제 (service_role 키로 RLS 우회)
    print("=== 기존 데이터 삭제 ===")
    for month in TARGET_MONTHS:
        # 삭제 전 카운트
        cnt_rows = fetch_all(f"/finance?month=eq.{month}&select=id")
        status, body = api("DELETE", f"/finance?month=eq.{month}")
        # 삭제 후 카운트
        after = fetch_all(f"/finance?month=eq.{month}&select=id")
        print(f"  {month}: {len(cnt_rows)} -> {len(after)} rows (DELETE status={status})")
    print()

    # xlsx 파싱
    print("\n=== xlsx 파싱 ===")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    all_rows = []
    for sheet_name, month in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"시트 없음: {sheet_name}")
            continue
        rows = parse_sheet(wb[sheet_name], month)
        all_rows.extend(rows)

    def do_insert(rows, label):
        BATCH = 200
        for i in range(0, len(rows), BATCH):
            chunk = rows[i:i+BATCH]
            status, body = api("POST", "/finance?on_conflict=id", chunk)
            if status >= 400:
                print(f"  INSERT ERROR {status}: {body[:300]}")
            else:
                print(f"  {label} {min(i+BATCH, len(rows))}/{len(rows)}\r", end="")
        print()

    # Pushbullet 복원 (별도 삽입)
    if pb_rows:
        print(f"\n=== Pushbullet 복원 ({len(pb_rows)}건) ===")
        do_insert(pb_rows, "pb")

    # xlsx 삽입
    print(f"\n=== xlsx 삽입 ({len(all_rows)}건) ===")
    do_insert(all_rows, "xlsx")
    print(f"DONE: pb {len(pb_rows)} + xlsx {len(all_rows)} = {len(pb_rows)+len(all_rows)} rows total")

if __name__ == "__main__":
    main()
