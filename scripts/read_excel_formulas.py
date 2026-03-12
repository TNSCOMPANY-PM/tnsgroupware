# Extract sheet names, formulas and sample values from the TNS revenue Excel file
import openpyxl
from pathlib import Path

path = Path(r"c:\Users\user1\Downloads\(주)티앤에스컴퍼니 매출 통계 24.01~ (1).xlsx")
if not path.exists():
    print("File not found:", path)
    exit(1)

wb = openpyxl.load_workbook(path, data_only=False)
print("=== SHEETS ===")
for name in wb.sheetnames:
    print(name)

for sheet_name in wb.sheetnames[:5]:  # first 5 sheets
    ws = wb[sheet_name]
    print("\n=== SHEET:", sheet_name, "===")
    formulas_found = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 20)):
        for cell in row:
            if cell.value is not None:
                val = cell.value
                if isinstance(val, str) and val.strip().startswith("="):
                    formulas_found.append((cell.coordinate, val))
                elif isinstance(val, (int, float)) and (cell.row <= 30 or "잔고" in str(ws.cell(1, cell.column).value or "") or "매출" in str(ws.cell(1, cell.column).value or "")):
                    pass  # skip raw numbers for brevity
    for coord, formula in formulas_found[:50]:
        print(coord, ":", formula[:120])
    if not formulas_found:
        # try data_only to see values
        ws2 = openpyxl.load_workbook(path, data_only=True)[sheet_name]
        for row in range(1, min(21, ws.max_row + 1)):
            row_vals = []
            for col in range(1, min(12, ws.max_column + 1)):
                v = ws2.cell(row, col).value
                if v is not None:
                    row_vals.append(f"{openpyxl.utils.get_column_letter(col)}{row}={v}")
            if row_vals:
                print("Row", row, ":", " | ".join(row_vals[:8]))

wb.close()
